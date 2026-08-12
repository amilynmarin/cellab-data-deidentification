from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .errors import InputDataError
from .models import ToolSchema


@dataclass
class SourceMetadata:
    selected_excel_sheet: str | None = None
    variable_labels: dict[str, str] = field(default_factory=dict)
    value_labels: dict[str, dict[str, str]] = field(default_factory=dict)
    source_formats: dict[str, str] = field(default_factory=dict)
    source_missing_codes: dict[str, list[Any]] = field(default_factory=dict)


@dataclass
class LoadedData:
    path: Path
    dataframe: pd.DataFrame
    metadata: SourceMetadata
    sha256: str


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _string_dtype_columns(schema: ToolSchema) -> dict[str, str]:
    columns: set[str] = set()
    for rule in schema.columns:
        if rule.validation.type == "string" or rule.action == "tokenize":
            columns.add(rule.source)
        for source in rule.token_sources:
            columns.add(source.column)
        columns.update(rule.token_context)
    columns.update(source.column for source in schema.shift.participant_sources)
    columns.update(source.column for source in schema.shift.dyad_sources)
    for group in schema.timestamp_groups:
        columns.update(group.anchor_columns)
    return {column: "string" for column in columns}


def _csv_header(path: Path, encoding: str = "utf-8-sig") -> list[str]:
    try:
        with path.open("r", encoding=encoding, newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader)
    except StopIteration as exc:
        raise InputDataError("CSV input is empty.") from exc
    except (OSError, UnicodeError, csv.Error) as exc:
        raise InputDataError("Unable to read the CSV header.") from exc
    names = [str(value) for value in header]
    if any(name == "" for name in names):
        raise InputDataError("CSV input contains a blank column name.")
    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        raise InputDataError(f"Input contains duplicate column names: {duplicates}")
    return names


def _sheet_has_values(sheet: Any) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(value is not None and value != "" for value in row):
            return True
    return False


def _populated_excel_sheets(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return [sheet.title for sheet in workbook.worksheets if _sheet_has_values(sheet)]
    finally:
        workbook.close()


def _read_selected_excel(path: Path, selected: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook[selected]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    while rows and all(value is None or value == "" for value in rows[-1]):
        rows.pop()
    if not rows:
        raise InputDataError("Selected Excel sheet is empty.")
    last_column = 0
    for row in rows:
        for index, value in enumerate(row, start=1):
            if value is not None and value != "":
                last_column = max(last_column, index)
    rows = [row[:last_column] + [None] * max(0, last_column - len(row)) for row in rows]
    header = ["" if value is None else str(value) for value in rows[0]]
    if any(name == "" for name in header):
        raise InputDataError("Excel input contains a blank column name in the populated range.")
    duplicates = sorted({name for name in header if header.count(name) > 1})
    if duplicates:
        raise InputDataError(f"Input contains duplicate column names: {duplicates}")
    return pd.DataFrame(rows[1:], columns=header, dtype="object")


def _read_excel(path: Path, schema: ToolSchema, sheet_override: str | None) -> tuple[pd.DataFrame, SourceMetadata]:
    populated = _populated_excel_sheets(path)
    if not populated:
        raise InputDataError("Excel workbook contains no populated sheets.")
    selected = sheet_override or schema.input.excel_sheet
    if selected is None:
        if len(populated) != 1:
            raise InputDataError(
                "Excel workbook contains multiple populated sheets; select one explicitly with --sheet or input.excel_sheet."
            )
        selected = populated[0]
    if selected not in populated:
        raise InputDataError("The selected Excel sheet is absent or empty.")
    return _read_selected_excel(path, selected), SourceMetadata(selected_excel_sheet=selected)


def _normalize_value_labels(raw: dict[Any, dict[Any, Any]] | None) -> dict[str, dict[str, str]]:
    """Normalize source metadata into JSON-safe, stable string mappings."""
    normalized: dict[str, dict[str, str]] = {}
    for variable, labels in (raw or {}).items():
        normalized[str(variable)] = {str(code): str(label) for code, label in labels.items()}
    return normalized


def _read_stata(path: Path) -> tuple[pd.DataFrame, SourceMetadata]:
    try:
        with pd.io.stata.StataReader(
            path,
            convert_categoricals=False,
            preserve_dtypes=True,
            convert_missing=False,
        ) as reader:
            frame = reader.read()
            variable_labels = reader.variable_labels()
            value_labels = reader.value_labels()
    except Exception as exc:
        raise InputDataError(f"Unable to read Stata input: {exc.__class__.__name__}.") from exc
    metadata = SourceMetadata(
        variable_labels={str(key): str(value) for key, value in variable_labels.items() if value},
        value_labels=_normalize_value_labels(value_labels),
    )
    return frame, metadata


def _read_spss(path: Path) -> tuple[pd.DataFrame, SourceMetadata]:
    try:
        import pyreadstat  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise InputDataError(
            "SPSS .sav support requires pyreadstat. Install the project with its declared runtime dependencies."
        ) from exc
    try:
        frame, meta = pyreadstat.read_sav(
            str(path),
            apply_value_formats=False,
            formats_as_category=False,
            user_missing=False,
        )
    except Exception as exc:  # pragma: no cover
        raise InputDataError(f"Unable to read SPSS input: {exc.__class__.__name__}.") from exc
    metadata = SourceMetadata(
        variable_labels={str(key): str(value) for key, value in (getattr(meta, "column_names_to_labels", {}) or {}).items() if value},
        value_labels=_normalize_value_labels(getattr(meta, "variable_value_labels", {}) or {}),
        source_formats={str(key): str(value) for key, value in (getattr(meta, "original_variable_types", {}) or {}).items()},
    )
    return frame, metadata


def load_input(
    path: str | Path,
    schema: ToolSchema,
    *,
    sheet_override: str | None = None,
) -> LoadedData:
    input_path = Path(path).expanduser().resolve()
    if not input_path.is_file():
        raise InputDataError("Input file does not exist or is not a regular file.")
    suffix = input_path.suffix.lower()
    metadata = SourceMetadata()
    try:
        if suffix == ".csv":
            _csv_header(input_path)
            frame = pd.read_csv(
                input_path,
                dtype=str,
                low_memory=False,
                keep_default_na=False,
                na_filter=False,
                skip_blank_lines=False,
                encoding="utf-8-sig",
            )
        elif suffix == ".xlsx":
            frame, metadata = _read_excel(input_path, schema, sheet_override)
        elif suffix == ".dta":
            frame, metadata = _read_stata(input_path)
        elif suffix == ".sav":
            frame, metadata = _read_spss(input_path)
        elif suffix in {".sas7bdat", ".xpt"}:
            raise InputDataError("SAS files are outside the v0.1 input scope.")
        else:
            raise InputDataError("Supported input formats are .csv, .xlsx, .sav, and .dta.")
    except InputDataError:
        raise
    except Exception as exc:
        raise InputDataError(f"Unable to read input data: {exc.__class__.__name__}.") from exc
    frame.columns = [str(column) for column in frame.columns]
    if frame.columns.duplicated().any():
        duplicates = sorted(set(frame.columns[frame.columns.duplicated()].astype(str)))
        raise InputDataError(f"Input contains duplicate column names: {duplicates}")
    frame = frame.reset_index(drop=True)
    metadata.source_formats = {
        **{column: str(dtype) for column, dtype in frame.dtypes.items()},
        **metadata.source_formats,
    }
    return LoadedData(
        path=input_path,
        dataframe=frame,
        metadata=metadata,
        sha256=sha256_file(input_path),
    )
